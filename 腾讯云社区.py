import re
import requests
import xlwt
from openpyxl import Workbook
from bs4 import BeautifulSoup
from time import sleep
import random
import linecache
url_base = 'https://cloud.tencent.com/developer/article/'
headers = { 
            'cache-control':'no-cache',
            'accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8',
            'accept-encoding' : 'gzip, deflate, br',
            'accept-language' : 'zh-CN,zh;q=0.9',
            'cookie' : 'qcloud_uid=be57c3d2723f8dcf1fd48b1175bde70d; tencent_uid=f890f9e77266b46d9b8594b7ac14a8c0; _ga=GA1.2.2136348531.1531808120; pgv_pvi=9739491328; pt2gguin=o0481344077; language=zh; qcloud_from=qcloud.baidu.seo-1532924806740; lastLoginType=qq; intl=; qcloud_visitId=06c7c3088db367a47745bf33b98335e7; _gat=1; pgv_si=s6490221568',
            'pragma' : 'no-cache',
            'referer' : 'https://cloud.tencent.com/developer',
            'upgrade-insecure-requests' : '1',
            'user-agent' : 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/68.0.3440.75 Safari/537.36',
            }
id_start = 1004664
id_end = 1004714
bc = 50
data = []

def data_mysql(data):
	'''
	Save article to mysql.
	'''

def data_save2(data,l,r):
	'''
	save article to file by xlwt
	'''
	book = xlwt.Workbook(encoding='utf-8', style_compression=0)
	sheet = book.add_sheet('云社区', cell_overwrite_ok=True)
	keys =list(data[0].keys());
	for i in range(len(keys)):
		sheet.write(0, i, keys[i])

	for i in range(len(data)):
		for j in range(len(keys)):
			sheet.write(i+1, j, data[i][keys[j]])
	book.save(r'E:\test'+str(l)+'-'+str(r)+'.xls')

def data_save(data,l,r):
	'''
	save article to file by openpyxl
	'''
	filename = r'E:\test'+str(l)+'-'+str(r)+'.xlsx'
	book = Workbook()
	books = book.active
	keys =list(data[0].keys());
	for i in range(len(data)+1):
		for j in range(len(keys)):
			if i == 0:
				books.cell(row=i+1, column=j+1, value = keys[j])
			else:
				books.cell(row=i+1, column=j+1, value = data[i-1][keys[j]])

	book.save(filename)
	print(filename,'已保存！\n')

def image2source(text):
	res1 = re.sub(r'(<figure><div class="image-block"><span class="lazy-image-holder" dataurl=")(\S*)("></span></div></figure>)', r'<a href ="\2"><img src="\2"/></a>', text)
	res = re.sub(r'\xa0', r'&nbsp;', res1)
	#res = ILLEGAL_CHARACTERS_RE.sub(r'', res)
	return res


def source2soup(url):
	the_line = 'http://'+linecache.getline(r'D:\cc\软件\ip.txt', random.randint(1,898))
	proxies = {
		'http':the_line.strip()
	}
	try:
		respose = requests.get(url,headers = headers, proxies = proxies, timeout = 0.5)

	except Timeout:
		print('Time Out:', url)
		return
	global data
	if respose.status_code == requests.codes.ok:
		respose.encoding = 'utf-8'
		Soup = BeautifulSoup(respose.text, 'html.parser')
		author_name = Soup.select('.author-name')
		column_name = Soup.select('.column-name')
		article_time = Soup.select('.col-article-time > span > time')
		article_title = Soup.select('.col-article-title')
		article = Soup.select('.J-articleContent')
		try:
			data.append({
				'article_url' : url,
				'author_name' : author_name[0].text,
				'author_url' : 'https://cloud.tencent.com/developer' + author_name[0]['href'],
				'column_name' : column_name[0].text,
				'column_url' : 'https://cloud.tencent.com/developer' + column_name[0]['href'],
				'article_title' : article_title[0].text,
				'article_time' : article_time[0]['datetime'],
				'article_content' : image2source(str(article[0]))
			})
		except IndexError:
			print('IndexError,sleep')
			sleep(0.5)
			data.append({
			'article_url' : url,
			'author_name' : 'Failed',
			'author_url' : 'Failed',
			'column_name' : 'Failed',
			'column_url' : 'Failed',
			'article_title' : 'Failed',
			'article_time' : 'Failed',
			'article_content' : 'Failed'
		})
			return
	else:
		data.append({
			'article_url' : url,
			'author_name' : '404 NOT FOUND',
			'author_url' : '404 NOT FOUND',
			'column_name' : '404 NOT FOUND',
			'column_url' : '404 NOT FOUND',
			'article_title' : '404 NOT FOUND',
			'article_time' : '404 NOT FOUND',
			'article_content' : '404 NOT FOUND'
		})
		sleep(0.2)

def main():
	for start in range(id_start, id_end, bc):
		if id_end-start>=bc:
			for article_id in range(start, start+bc):
				url = url_base + str(article_id)
				source2soup(url)
			data_save(data, start, start+bc)
			del data[:]
			sleep(0.1)
		else:
			for article_id in range(start, id_end):
				url = url_base + str(article_id)
				source2soup(url)
			data_save(data, start, id_end)
			del data[:]

if __name__ == '__main__':
	main()